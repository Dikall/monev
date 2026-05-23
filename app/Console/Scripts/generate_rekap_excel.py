"""
generate_rekap_excel.py
========================
Dipanggil oleh AdminController::exportExcelFormatted().
Argumen CLI:
  1. Path ke file JSON input  (data rekap nilai dari Laravel)
  2. Path ke file XLSX output (yang akan ditulis dan dikirim kembali ke PHP)

Dependensi:  openpyxl  (pip install openpyxl)
"""

import sys
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------
RED_FILL   = PatternFill("solid", fgColor="C0392B")
GREY_FILL  = PatternFill("solid", fgColor="F2F2F2")
GREEN_FILL = PatternFill("solid", fgColor="E8F5E9")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")

THIN   = Side(style="thin", color="BDBDBD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WHITE_BOLD   = Font(name="Calibri", bold=True,  color="FFFFFF", size=10)
BLACK_BOLD   = Font(name="Calibri", bold=True,  color="000000", size=10)
BLACK_NORMAL = Font(name="Calibri", bold=False, color="000000", size=10)
GREY_NORMAL  = Font(name="Calibri", bold=False, color="BDBDBD", size=10)
RED_BOLD     = Font(name="Calibri", bold=True,  color="C0392B", size=10)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def sc(cell, font=None, fill=None, alignment=None, border=BORDER):
    """Apply styles to a cell."""
    if font:      cell.font      = font
    if fill:      cell.fill      = fill
    if alignment: cell.alignment = alignment
    if border is not None:
        cell.border = border


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    if len(sys.argv) < 3:
        print("Usage: generate_rekap_excel.py <input.json> <output.xlsx>",
              file=sys.stderr)
        sys.exit(1)

    json_path = sys.argv[1]
    xlsx_path = sys.argv[2]

    with open(json_path, "r", encoding="utf-8-sig") as f:
        data = json.load(f)

    verifikator_name = data.get("verifikator_name", "-")
    tahun            = data.get("tahun", "-")
    tanggal_cetak    = data.get("tanggal_cetak", "-")
    indikators       = data.get("indikators", [])
    rows             = data.get("rows", [])

    wb = Workbook()
    ws = wb.active
    ws.title = f"Rekap Nilai {tahun}"

    # Total columns: No | Kategori | Nama Badan | Responden | [ind...] | Total | Presentasi | Score
    total_cols = 4 + len(indikators) + 3

    # -------------------------------------------------------------------
    # Row 1-2: Judul
    # -------------------------------------------------------------------
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c = ws.cell(1, 1, f"REKAP NILAI KUESIONER \u2013 TAHUN {tahun}")
    sc(c, font=Font(name="Calibri", bold=True, color="FFFFFF", size=14),
       fill=RED_FILL, alignment=CENTER, border=None)
    ws.row_dimensions[1].height = 32

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    c = ws.cell(2, 1,
                f"Verifikator: {verifikator_name}   |   Tanggal Cetak: {tanggal_cetak}")
    sc(c, font=Font(name="Calibri", bold=False, color="FFFFFF", size=10),
       fill=RED_FILL, alignment=CENTER, border=None)
    ws.row_dimensions[2].height = 18

    # blank separator
    ws.row_dimensions[3].height = 6

    # -------------------------------------------------------------------
    # Row 4: Header
    # -------------------------------------------------------------------
    HDR = 4
    ws.row_dimensions[HDR].height = 40

    fixed_headers = ["No", "Kategori", "Nama Badan Publik", "Nama Responden"]
    ind_headers   = [
        f"{i['nama_indikator']}\n(Bobot {i['bobot']})"
        for i in indikators
    ]
    tail_headers  = ["Total Kuesioner", "Nilai Presentasi", "Total Score"]

    for col_idx, header in enumerate(fixed_headers + ind_headers + tail_headers, start=1):
        c = ws.cell(HDR, col_idx, header)
        sc(c, font=WHITE_BOLD, fill=RED_FILL, alignment=CENTER)

    # -------------------------------------------------------------------
    # Data rows
    # -------------------------------------------------------------------
    for row_data in rows:
        r = HDR + row_data["no"]
        ws.row_dimensions[r].height = 18
        bg = GREEN_FILL if row_data.get("is_submitted") else WHITE_FILL

        col = 1

        def wcell(val, font=BLACK_NORMAL, align=CENTER):
            c = ws.cell(r, col, val)
            sc(c, font=font, fill=bg, alignment=align)
            return c

        # No
        c = ws.cell(r, col, row_data["no"])
        sc(c, font=BLACK_NORMAL, fill=bg, alignment=CENTER)
        col += 1

        # Kategori
        c = ws.cell(r, col, row_data["kategori"])
        sc(c, font=BLACK_NORMAL, fill=bg, alignment=CENTER)
        col += 1

        # Nama Badan Publik
        c = ws.cell(r, col, row_data["nama_badan"])
        sc(c, font=BLACK_NORMAL, fill=bg, alignment=LEFT)
        col += 1

        # Nama Responden
        c = ws.cell(r, col, row_data["nama_responden"])
        sc(c, font=BLACK_NORMAL, fill=bg, alignment=LEFT)
        col += 1

        # Nilai per Indikator
        nilai_map = {str(k): v for k, v in row_data.get("nilai_per_indikator", {}).items()}
        for ind in indikators:
            val = nilai_map.get(str(ind["id"]))
            if val is None:
                c = ws.cell(r, col, "\u2013")
                sc(c, font=GREY_NORMAL, fill=bg, alignment=CENTER)
            else:
                c = ws.cell(r, col, float(val))
                sc(c, font=BLACK_NORMAL, fill=bg, alignment=CENTER)
                c.number_format = "0.00"
            col += 1

        # Total Kuesioner
        c = ws.cell(r, col, float(row_data.get("total_kuesioner", 0)))
        sc(c, font=BLACK_BOLD, fill=bg, alignment=CENTER)
        c.number_format = "0.00"
        col += 1

        # Nilai Presentasi
        np_val = row_data.get("nilai_presentasi")
        if np_val is None:
            c = ws.cell(r, col, "\u2014")
            sc(c, font=GREY_NORMAL, fill=bg, alignment=CENTER)
        else:
            c = ws.cell(r, col, float(np_val))
            sc(c, font=BLACK_NORMAL, fill=bg, alignment=CENTER)
            c.number_format = "0.00"
        col += 1

        # Total Score
        ts_val = row_data.get("total_score")
        if ts_val is None:
            c = ws.cell(r, col, "\u2014")
            sc(c, font=GREY_NORMAL, fill=bg, alignment=CENTER)
        else:
            c = ws.cell(r, col, float(ts_val))
            sc(c, font=RED_BOLD, fill=bg, alignment=CENTER)
            c.number_format = "0.00"

    # -------------------------------------------------------------------
    # Keterangan footer
    # -------------------------------------------------------------------
    note_row = HDR + len(rows) + 2
    ws.merge_cells(start_row=note_row, start_column=1,
                   end_row=note_row, end_column=total_cols)
    c = ws.cell(note_row, 1,
                "Keterangan: Total Score = (Total Kuesioner \u00d7 70%) + (Nilai Presentasi \u00d7 30%)"
                "   |   Baris hijau = Sudah Submit")
    sc(c, font=Font(name="Calibri", italic=True, color="555555", size=9),
       fill=None, alignment=LEFT, border=None)

    # -------------------------------------------------------------------
    # Column widths & freeze
    # -------------------------------------------------------------------
    widths = [5, 20, 36, 26] + [15] * len(indikators) + [16, 16, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = ws.cell(HDR + 1, 5)  # freeze header + 4 fixed cols

    wb.save(xlsx_path)
    print(f"OK: {xlsx_path}")


if __name__ == "__main__":
    main()
