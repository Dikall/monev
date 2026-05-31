"""
generate_list_akun_excel.py
Generates a styled .xlsx recap file for admin list-akun page.
Usage: python generate_list_akun_excel.py <input.json> <output.xlsx>
"""

import sys
import json
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── Colour palette ──────────────────────────────────────────────────────────
RED_DARK   = "9B1C1C"   # header rows
RED_LIGHT  = "FEE2E2"   # sheet-2 section header
GREEN_DARK = "166534"   # sheet-1 section header
GREEN_LIGHT= "DCFCE7"   # sheet-1 zebra
AMBER_LIGHT= "FEF3C7"   # sheet-2 zebra
WHITE      = "FFFFFF"
GRAY_LIGHT = "F9FAFB"
GRAY_TEXT  = "6B7280"

def thin_border(color="D1D5DB"):
    s = Side(border_style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_font(color=WHITE, bold=True, size=10):
    return Font(name="Calibri", bold=bold, color=color, size=size)

def cell_font(bold=False, size=10, color="111827"):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── Build one sheet ──────────────────────────────────────────────────────────
def build_sheet(ws, rows, indikators, sheet_title, is_sudah_submit, kategori, tahun, tanggal_cetak, verifikator_name):
    """Populate a single worksheet."""

    ind_count = len(indikators)
    # Cols: No | Nama Badan | Nama Responden | Status? | ind1..N | Nilai Presentasi | Total | Total Score
    has_status_col = not is_sudah_submit   # "tidak mengisi" sheet has a Status column
    STATUS_OFFSET  = 1 if has_status_col else 0
    TOTAL_COLS     = 3 + STATUS_OFFSET + ind_count + 3   # No+Nama+Responden+(Status)+inds+NilPres+Total+Score
    last_col_letter = get_column_letter(TOTAL_COLS)

    # ── Title block ──────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = "REKAP NILAI KUESIONER"
    ws["A1"].font      = Font(name="Calibri", bold=True, size=14, color=WHITE)
    ws["A1"].fill      = fill(RED_DARK)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{last_col_letter}2")
    ws["A2"] = f"Kategori: {kategori}  |  Tahun: {tahun}  |  Verifikator: {verifikator_name}  |  Tanggal Cetak: {tanggal_cetak}"
    ws["A2"].font      = Font(name="Calibri", italic=True, size=9, color=WHITE)
    ws["A2"].fill      = fill(RED_DARK)
    ws["A2"].alignment = center()
    ws.row_dimensions[2].height = 18

    ws.merge_cells(f"A3:{last_col_letter}3")
    ws["A3"] = sheet_title
    title_fill = fill(GREEN_DARK) if is_sudah_submit else fill("991B1B")
    ws["A3"].font      = Font(name="Calibri", bold=True, size=11, color=WHITE)
    ws["A3"].fill      = title_fill
    ws["A3"].alignment = center()
    ws.row_dimensions[3].height = 22

    # ── Column headers ───────────────────────────────────────────────────────
    HDR_ROW = 4
    col = 1

    def write_hdr(r, c, val, col_span=1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font      = hdr_font()
        cell.fill      = fill(RED_DARK)
        cell.alignment = center()
        cell.border    = thin_border(RED_DARK)
        if col_span > 1:
            ws.merge_cells(start_row=r, start_column=c,
                           end_row=r, end_column=c + col_span - 1)

    write_hdr(HDR_ROW, col, "No");              col += 1
    write_hdr(HDR_ROW, col, "Nama Badan Publik"); col += 1
    write_hdr(HDR_ROW, col, "Nama Responden");    col += 1

    if has_status_col:
        write_hdr(HDR_ROW, col, "Status");       col += 1

    for ind in indikators:
        label = f"{ind['nama_indikator']}\n(Bobot: {ind['bobot']})"
        write_hdr(HDR_ROW, col, label);          col += 1

    write_hdr(HDR_ROW, col, "Nilai Presentasi"); col += 1
    write_hdr(HDR_ROW, col, "Total Kuesioner");  col += 1
    write_hdr(HDR_ROW, col, "Total Score");      col += 1

    ws.row_dimensions[HDR_ROW].height = 36

    # ── Data rows ────────────────────────────────────────────────────────────
    zebra_fill = fill(GREEN_LIGHT) if is_sudah_submit else fill(AMBER_LIGHT)

    for idx, row in enumerate(rows):
        r = HDR_ROW + 1 + idx
        alt = (idx % 2 == 1)
        base_fill = zebra_fill if alt else fill(WHITE)

        col = 1
        def wr(val, bold=False, color="111827", h_align="center"):
            c = ws.cell(row=r, column=col, value=val)
            c.font      = cell_font(bold=bold, color=color)
            c.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=True)
            c.fill      = base_fill
            c.border    = thin_border()
            return c

        wr(row["no"]);                                                          col += 1
        wr(row["nama_badan"],    bold=True, h_align="left");                    col += 1
        wr(row["nama_responden"], h_align="left");                              col += 1

        if has_status_col:
            status = row.get("status", "-")
            sc = ws.cell(row=r, column=col, value=status)
            sc.font      = cell_font(bold=True, color="92400E" if "Sedang" in status else "374151")
            sc.alignment = center()
            sc.fill      = fill("FEF9C3") if "Sedang" in status else fill("F3F4F6")
            sc.border    = thin_border()
            col += 1

        npi = row.get("nilai_per_indikator", {})
        for ind in indikators:
            val = npi.get(str(ind["id"]), npi.get(ind["id"], 0))
            c = ws.cell(row=r, column=col, value=val if val != 0 else 0)
            c.font      = cell_font(bold=True, color="1D4ED8" if val and val > 0 else GRAY_TEXT)
            c.alignment = center()
            c.fill      = base_fill
            c.border    = thin_border()
            col += 1

        # Nilai Presentasi
        np_val = row.get("nilai_presentasi")
        c = ws.cell(row=r, column=col,
                    value=np_val if np_val is not None else "Belum diisi")
        c.font      = cell_font(bold=(np_val is not None),
                                color="1D4ED8" if np_val is not None else GRAY_TEXT)
        c.alignment = center()
        c.fill      = base_fill
        c.border    = thin_border()
        col += 1

        # Total Kuesioner
        tk = ws.cell(row=r, column=col, value=row.get("total_kuesioner", 0))
        tk.font      = cell_font(bold=True, color="111827")
        tk.alignment = center()
        tk.fill      = base_fill
        tk.border    = thin_border()
        col += 1

        # Total Score
        ts_val = row.get("total_score")
        ts = ws.cell(row=r, column=col,
                     value=ts_val if ts_val is not None else "-")
        score_color = ("166534" if is_sudah_submit else "991B1B") if ts_val is not None else GRAY_TEXT
        ts.font      = cell_font(bold=True, color=score_color)
        ts.alignment = center()
        ts.fill      = base_fill
        ts.border    = thin_border()
        col += 1

        ws.row_dimensions[r].height = 18

    # ── Legend row ───────────────────────────────────────────────────────────
    legend_r = HDR_ROW + 1 + len(rows) + 1
    ws.merge_cells(f"A{legend_r}:{last_col_letter}{legend_r}")
    ws[f"A{legend_r}"] = "Keterangan: Total Score = (Total Kuesioner × 70%) + (Nilai Presentasi × 30%)"
    ws[f"A{legend_r}"].font      = Font(name="Calibri", italic=True, size=9, color=GRAY_TEXT)
    ws[f"A{legend_r}"].alignment = left()
    ws.row_dimensions[legend_r].height = 16

    # ── Column widths ────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 6    # No
    ws.column_dimensions["B"].width = 32   # Nama Badan
    ws.column_dimensions["C"].width = 24   # Nama Responden

    col_offset = 4
    if has_status_col:
        ws.column_dimensions[get_column_letter(col_offset)].width = 18
        col_offset += 1

    for i in range(ind_count):
        ws.column_dimensions[get_column_letter(col_offset + i)].width = 14

    after_ind = col_offset + ind_count
    ws.column_dimensions[get_column_letter(after_ind)].width     = 16  # Nilai Presentasi
    ws.column_dimensions[get_column_letter(after_ind + 1)].width = 14  # Total
    ws.column_dimensions[get_column_letter(after_ind + 2)].width = 14  # Score

    # Freeze header rows
    ws.freeze_panes = ws.cell(row=HDR_ROW + 1, column=1)


# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 3:
        print("Usage: python generate_list_akun_excel.py <input.json> <output.xlsx>")
        sys.exit(1)

    json_path = sys.argv[1]
    xlsx_path = sys.argv[2]

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    kategori         = data.get("kategori", "")
    tahun            = data.get("tahun", "")
    tanggal_cetak    = data.get("tanggal_cetak", "")
    verifikator_name = data.get("verifikator_name", "")
    export_type      = data.get("export_type", "all")
    indikators       = data.get("indikators", [])
    rows_mengisi     = data.get("rows_mengisi", [])
    rows_tidak       = data.get("rows_tidak", [])

    wb = Workbook()
    
    if export_type == "mengisi" or export_type == "all":
        # Sheet 1: Sudah Mengisi
        ws1 = wb.active
        ws1.title = "Sudah Submit"
        build_sheet(
            ws1, rows_mengisi, indikators,
            f"Rekapan Nilai — Sudah Submit ({len(rows_mengisi)} Badan Publik)",
            is_sudah_submit=True,
            kategori=kategori, tahun=tahun,
            tanggal_cetak=tanggal_cetak, verifikator_name=verifikator_name
        )

    if export_type == "tidak" or export_type == "all":
        # Sheet 2: Belum Mengisi
        if export_type == "tidak":
            ws2 = wb.active
            ws2.title = "Belum Submit"
        else:
            ws2 = wb.create_sheet(title="Belum Submit")
            
        build_sheet(
            ws2, rows_tidak, indikators,
            f"Rekapan Nilai — Tidak Mengisi Kuesioner ({len(rows_tidak)} Badan Publik)",
            is_sudah_submit=False,
            kategori=kategori, tahun=tahun,
            tanggal_cetak=tanggal_cetak, verifikator_name=verifikator_name
        )

    wb.save(xlsx_path)
    print(f"OK: {xlsx_path}")

if __name__ == "__main__":
    main()
